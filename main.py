import os
import random

import openpyxl

file_list = os.listdir(os.getcwd())
for file in file_list:
    if '.xls' in file:
        file_name = file


def random_execl(filename, sheetname=None):
    wb = openpyxl.load_workbook(filename)
    if sheetname is None:
        sheetname = wb.sheetnames[0]
    # 指定要读取内容的sheet
    sheetobj = wb[sheetname]
    # 读取所有的内容
    # 获取最大行
    rows = sheetobj.max_row
    # 获取最大列
    columns = sheetobj.max_column
    for column in range(1, columns + 1):
        if '收缩压' in sheetobj.cell(1, column).value or '舒张压' in sheetobj.cell(1, column).value:
            column_change = column
            for row in range(2, rows + 1):
                bp = sheetobj.cell(row, column_change).value
                sheetobj.cell(row, column_change).value = int(bp) + random.randint(-15, 15)
        elif sheetobj.cell(1, column).value == '血糖':
            column_change = column
            for row in range(2, rows + 1):
                bp = sheetobj.cell(row, column_change).value
                sheetobj.cell(row, column_change).value = format(float(bp) + random.uniform(-1.2, 1.2),'.1f')
    wb.save('E_'+filename)

random_execl(file_name)
